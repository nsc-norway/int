import sys
import os
import base64
import json
import io
import requests
from openpyxl import load_workbook

SERVER="https://portal.sequencing.uio.no"
MAX_SAMPLES = 10000

class WorksheetFormatException(Exception):
    pass


def validate_workbook(wb, expected_num_samples):
    """Throws exception if the format is incorrect.

    Returns errors if the data entry is incorrect."""
    ws = wb.worksheets[0]

    headers = [
            "Number",
            "Plate",
            "Sample name",
            "Conc.",
            "A260/280",
            "A260/230",
            "Volume provided",
            "Total DNA / RNA",
            "Index name",
            "Index Seq",
            "Primers, Linkers or RE",
            "Approx no. Reads"
            ]

    col_of = {}

    for i,header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i).lower()
        if cell.startswith(header.lower()):
            col_of[header] = i
        else:
            raise WorksheetFormatException()

    plates = {}
    for row in range(2, MAX_SAMPLES):
        number = ws.cell(row=row, column=col_of['Number'])
        name = ws.cell(row=row, column=col_of['Sample name'])

        
    

    return []


def validate_sample_table(order_id, apikey):
    order = requests.get(
            '{SERVER}/api/v1/order/{order_id}'.format(SERVER=SERVER, order_id=order_id),
            headers = {'X-OrderPortal-API-key': apikey}
            ).json()
    files = order['files']
    
    errors = []
    found_sample_table = False

    for f in files:
        if f['filename'].endswith(".xls"):
            data = requests.get(
                    '{SERVER}/orders/{order_id}/file/{filename}'.format(
                        SERVER=SERVER, order_id=order_id, filename=f['filename']
                        )
                    ).content
            io_data = BytesIO(data)
            try:
                wb = load_workbook(io_data)
            except:
                continue

            try:
                messages = validate_workbook(wb)
            except ValueError:
                continue

            if found_workbook:
                errors.append("Multiple sample tables found. Please add all your\
                        samples to a single sample table")

            found_workbook = True

            try:
                errors += validate_workbook(wb)
            except:
                errors.append("Unexpected error reading " + f['filename'])




def main():
    apikey = open(os.path.expanduser("~/portal-apikey")).read().strip()
    with open("testorder.json", "w") as output_file:
    #with io.StringIO() as output_file:
        write_order_package("d6b4c1d7ffae4d1e9e9065e4e7e4a97d", apikey, output_file)
        #print(output_file.getvalue())

if __name__ == "__main__":
    main()

